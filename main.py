import openpyxl

# Load the workbook to use in this script
inv_file = openpyxl.load_workbook("Inventory.xlsx")
product_list = inv_file["Sheet1"]
product_list.cell(row=1, column=5, value="Total Value")     # Create new column with header name

products_per_supplier = {}
total_value_per_supplier = {}
prod_under_10_inv = {}

# Start range iteration from row 2 since row 1 has headers
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value     # Creates key-value pair & extracts name from col4
    inventory = product_list.cell(product_row, 2).value         # Inventory for col 2
    price = product_list.cell(product_row, 3).value             # Price for col 3
    product_num = product_list.cell(product_row, 1).value       # Product number for col 1
    inv_total_per_prod = product_list.cell(product_row, 5)      # Create a new value & new column starting row 2

    # Calculation for number of products per supplier
    if supplier_name in products_per_supplier:                      # Check if supplier already exists in dictionary...
        current_num_products = products_per_supplier.get(supplier_name)     # If so, grab the current count total...
        products_per_supplier[supplier_name] = current_num_products + 1     # ...increase count by 1 and add to total
    else:
        print(f"Adding a new supplier '{supplier_name}'")
        products_per_supplier[supplier_name] = 1        # Creates new entry in dict for new supplier and sets count to 1

    # Calculate total value of inventory per supplier. Similar logic to above; instead of inc count by 1, add to total
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price     # Create new entry with product's total value

    # Find products with inventory less than 10
    if inventory < 10:
        prod_under_10_inv[int(product_num)] = int(inventory)

    # Add a total value for inventory and price for a unique product no.
    inv_total_per_prod.value = inventory * price

# Save new calculated inventory totals per product no. to a new file
inv_file.save("updated_inventory_totals.xlsx")

print(f"\nProduct No. with total inventory less than 10:\n{prod_under_10_inv}")
print(f"\nTotal number of product counts for each unique company:\n{products_per_supplier}")
print(f"\nTotal amounts of all inventory values for each unique company:\n{total_value_per_supplier}")
